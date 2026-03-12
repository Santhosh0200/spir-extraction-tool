"""
extraction/column_mapper.py
────────────────────────────
Dynamic Column Detection Engine
────────────────────────────────
Instead of hardcoding column numbers, this module scans a worksheet's header
rows for keyword patterns and returns a mapping of logical field names to
physical column indices.

Design goals:
  • Works with any SPIR variant regardless of column ordering
  • Tolerant of extra whitespace, line-breaks, and minor typos
  • Returns a confidence score so the caller can decide to trust or warn
  • Never raises — returns an empty mapping entry if a column is not found

Usage:
    mapper = ColumnMapper(ws, header_row=6)
    col_desc = mapper.get('DESCRIPTION')      # → int column index, or None
    col_sap  = mapper.get('SAP_NUMBER')       # → int column index, or None
    report   = mapper.report()                # → dict of all mappings + scores
"""
from __future__ import annotations
import re
import logging
from dataclasses import dataclass, field

log = logging.getLogger(__name__)


# ── Keyword catalogue ─────────────────────────────────────────────────────────
# Each logical field maps to a list of keyword patterns (priority order).
# The first keyword that matches wins. Matching is case-insensitive substring.

FIELD_KEYWORDS: dict[str, list[str]] = {
    # Core item fields
    "ITEM_NUMBER":          ["item number", "item no", "item#", "item num"],
    "DESCRIPTION":          ["description of parts", "description of part",
                             "description", "desc of part"],
    "DWG_NO":               ["dwg no", "drawing no", "dwg number", "pos'n no",
                             "posn no", "drawing number"],
    "MFR_PART_NO":          ["manufacturer part", "mfr part", "manuf part",
                             "part number", "part no"],
    "SUPPLIER_PART_NO":     ["supplier part", "sup part", "suppliers part"],
    "SUPPLIER_NAME":        ["supplier/ocm", "supplier name", "suppliers name",
                             "supplier", "ocm name"],
    "MATERIAL_SPEC":        ["material spec", "material spec'n"],
    "MATERIAL_CERT":        ["material cert", "inspection cert"],
    "CURRENCY":             ["currency"],
    "UNIT_PRICE":           ["unit price", "price"],
    "DELIVERY":             ["delivery time", "delivery"],
    "MIN_MAX":              ["min/max", "min max", "stock lvl"],
    "UOM":                  ["unit of measure", "uom"],
    "SAP_NUMBER":           ["sap number", "sap no", "sap#"],
    "CLASSIFICATION":       ["classification", "class of part"],

    # Quantity fields
    "QTY_IDENTICAL":        ["total no. of identical", "identical parts fitted",
                             "no. of identical", "qty identical"],
    "NO_OF_UNITS":          ["no. of units", "number of units", "eqpt qty",
                             "no of units"],

    # Tag / equipment fields
    "EQUIPMENT_TAG":        ["equipment tag", "equip tag", "equip't or tag",
                             "tag no", "tag number"],
    "MFR_MODEL":            ["mfr type or model", "mfr model", "manufacturer model",
                             "type or model", "model number"],
    "MFR_SERIAL":           ["mfr ser'l", "mfr serial", "serial no",
                             "serial number"],
    "MFR_NAME":             ["manufacturer name", "manufacturer:", "made by"],
    "SUPPLIER_COMPANY":     ["supplier company", "supplier/ocm name",
                             "vendor name"],

    # Spare type flags
    "INITIAL_SPARES":       ["initial spare", "initial spares"],
    "NORMAL_OP_SPARES":     ["normal operat", "normal oper", "normal spares"],
    "COMMISSIONING_SPARES": ["commissioning spare", "commissioning spares"],
    "LIFECYCLE_SPARES":     ["life cycle spare", "lifecycle spare"],

    # Recommendation columns
    "RECOM_BY_MFR_INIT":    ["recommended by .manufacturer", "recom.*manuf.*initial",
                             "initial.*recommended.*manuf"],
    "RECOM_BY_MFR_NORM":    ["recommended by .manufacturer", "recom.*manuf.*normal"],

    # Remarks / References
    "REMARKS":              ["remarks", "remark"],
    "ITEM_REF":             ["item ref", "item reference"],
}


def _normalise(text: str) -> str:
    """Lowercase, collapse whitespace, strip special chars for fuzzy matching."""
    t = str(text or "").lower()
    t = re.sub(r"[\r\n]+", " ", t)
    t = re.sub(r"\s+", " ", t)
    return t.strip()


@dataclass
class _Match:
    col_idx:    int
    col_letter: str
    raw_text:   str
    keyword:    str
    score:      float   # 0.0-1.0


@dataclass
class ColumnMapper:
    """
    Scan a worksheet for known column headers and produce a logical→physical map.

    Args:
        ws:           openpyxl worksheet object.
        header_rows:  Which rows to scan (default: rows 4–8, covering most SPIR layouts).
        max_col:      Scan up to this column index (0 = auto-detect from max_column).
    """
    ws:           object                        # openpyxl Worksheet
    header_rows:  list[int]  = field(default_factory=lambda: [4, 5, 6, 7, 8])
    max_col:      int        = 0

    # Internal state
    _mapping: dict[str, _Match | None] = field(default_factory=dict, init=False)
    _scanned: bool = field(default=False, init=False)

    # ── Public API ────────────────────────────────────────────────────────────

    def get(self, field_name: str) -> int | None:
        """
        Return the 1-based column index for a logical field, or None if not found.

        Example:
            mapper.get('SAP_NUMBER')  →  26
        """
        self._ensure_scanned()
        match = self._mapping.get(field_name)
        return match.col_idx if match else None

    def get_or_raise(self, field_name: str) -> int:
        col = self.get(field_name)
        if col is None:
            raise KeyError(f"Column '{field_name}' not found in worksheet '{self.ws.title}'")
        return col

    def report(self) -> dict[str, dict]:
        """
        Return a full mapping report — useful for API diagnostics and logging.

        Returns:
            {
                'DESCRIPTION': {'col': 9, 'letter': 'I', 'text': 'DESCRIPTION OF PARTS', 'score': 1.0},
                'SAP_NUMBER':  {'col': 26, 'letter': 'Z', 'text': 'SAP NUMBER', 'score': 1.0},
                'UNIT_PRICE':  None,
                ...
            }
        """
        self._ensure_scanned()
        out = {}
        for fname, match in self._mapping.items():
            if match is None:
                out[fname] = None
            else:
                out[fname] = {
                    "col":    match.col_idx,
                    "letter": match.col_letter,
                    "text":   match.raw_text,
                    "score":  match.score,
                }
        return out

    def coverage(self) -> float:
        """Return fraction of fields that were successfully mapped (0.0-1.0)."""
        self._ensure_scanned()
        found = sum(1 for m in self._mapping.values() if m is not None)
        return found / len(FIELD_KEYWORDS) if FIELD_KEYWORDS else 0.0

    # ── Internal ──────────────────────────────────────────────────────────────

    def _ensure_scanned(self) -> None:
        if not self._scanned:
            self._scan()
            self._scanned = True

    def _scan(self) -> None:
        """
        Walk every cell in the target rows and attempt keyword matching
        for each logical field.  First match wins (priority order).
        """
        ws = self.ws
        max_col = self.max_col or ws.max_column

        # Build a cell text table keyed by (row, col)
        cell_texts: dict[tuple[int, int], str] = {}
        for ri in self.header_rows:
            if ri > ws.max_row:
                continue
            for ci in range(1, max_col + 1):
                raw = ws.cell(ri, ci).value
                if raw is not None:
                    cell_texts[(ri, ci)] = _normalise(str(raw))

        # Initialise all fields to None
        mapping: dict[str, _Match | None] = {f: None for f in FIELD_KEYWORDS}

        # Match each field
        for field_name, keywords in FIELD_KEYWORDS.items():
            for keyword in keywords:
                kw_norm = _normalise(keyword)
                # Try regex if keyword contains special chars, otherwise substring
                use_regex = bool(re.search(r'[.*+?^${}()|[\]\\]', keyword))

                for (ri, ci), text in cell_texts.items():
                    hit = False
                    if use_regex:
                        try:
                            hit = bool(re.search(kw_norm, text))
                        except re.error:
                            hit = kw_norm in text
                    else:
                        hit = kw_norm in text

                    if hit:
                        from openpyxl.utils import get_column_letter
                        # Score: exact match = 1.0, substring = 0.8, regex = 0.7
                        score = 1.0 if text == kw_norm else (0.7 if use_regex else 0.8)
                        mapping[field_name] = _Match(
                            col_idx    = ci,
                            col_letter = get_column_letter(ci),
                            raw_text   = text,
                            keyword    = keyword,
                            score      = score,
                        )
                        break   # first matching keyword wins for this field

                if mapping[field_name] is not None:
                    break   # found — no need to try more keywords

        self._mapping = mapping
        found_count = sum(1 for m in mapping.values() if m is not None)
        log.debug(
            "ColumnMapper: sheet='%s'  found %d/%d fields",
            ws.title, found_count, len(FIELD_KEYWORDS)
        )


def map_sheet(ws, header_rows: list[int] | None = None) -> ColumnMapper:
    """
    Convenience factory — create and return a pre-scanned ColumnMapper.

    Args:
        ws:           openpyxl worksheet
        header_rows:  rows to scan (default covers rows 4-8)

    Returns:
        ColumnMapper (already scanned, ready to call .get())
    """
    rows = header_rows or [4, 5, 6, 7, 8]
    mapper = ColumnMapper(ws=ws, header_rows=rows)
    mapper._scan()
    mapper._scanned = True
    return mapper
