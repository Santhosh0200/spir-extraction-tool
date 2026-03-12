"""
extraction/spir_detector.py
───────────────────────────
Format detection and file validation layer.
Wraps the engine's detect_format() with richer validation and logging so
the API layer gets structured feedback before committing to extraction.
"""
from __future__ import annotations
import io
import logging

import openpyxl

from extraction.spir_engine import detect_format

log = logging.getLogger(__name__)

# Allowed MIME / extension combos
_ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}

# Human-readable format descriptions used in API responses
FORMAT_DESCRIPTIONS: dict[str, str] = {
    "FORMAT1": "Multi-Annexure SPIR (.xlsx)",
    "FORMAT2": "Single-Sheet, 1 Tag (.xlsm)",
    "FORMAT3": "Single-Sheet, Multi-Tag (.xlsm)",
    "FORMAT4": "Matrix SPIR + Single Continuation Sheet (.xlsx)",
    "FORMAT5": "Flag SPIR + Multiple Continuation Sheets (.xlsm)",
}


class ValidationError(Exception):
    """Raised when the uploaded file fails pre-processing validation."""


def validate_file(filename: str, content: bytes, max_mb: int = 100) -> None:
    """
    Validate filename extension and file size before touching openpyxl.

    Raises:
        ValidationError with a user-friendly message.
    """
    # Extension check
    lower = filename.lower()
    if not any(lower.endswith(ext) for ext in _ALLOWED_EXTENSIONS):
        raise ValidationError(
            f"Unsupported file type '{filename}'. "
            f"Please upload one of: {', '.join(_ALLOWED_EXTENSIONS)}"
        )

    # Size check
    size_mb = len(content) / (1024 * 1024)
    if size_mb > max_mb:
        raise ValidationError(
            f"File too large ({size_mb:.1f} MB). Maximum allowed: {max_mb} MB."
        )

    # Minimal structure check — can openpyxl open it at all?
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet_count = len(wb.sheetnames)
        wb.close()
        if sheet_count == 0:
            raise ValidationError("The workbook contains no sheets.")
    except ValidationError:
        raise
    except Exception as exc:
        raise ValidationError(f"Could not open workbook: {exc}") from exc


def detect(content: bytes) -> dict:
    """
    Open the workbook and return format metadata.

    Returns:
        {
            'format':      'FORMAT3',
            'description': 'Single-Sheet, Multi-Tag (.xlsm)',
            'sheets':      ['MAIN SHEET', 'Validation Lists'],
        }
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    fmt = detect_format(wb)
    sheets = wb.sheetnames
    wb.close()

    log.info("Format detected: %s  sheets=%s", fmt, sheets)
    return {
        "format":      fmt,
        "description": FORMAT_DESCRIPTIONS.get(fmt, fmt),
        "sheets":      sheets,
    }
