"""
tests/test_extraction.py
─────────────────────────
Tests for the core extraction engine and supporting modules.

Run:
    pytest tests/ -v
    pytest tests/ -v --tb=short -q     # quiet
"""
from __future__ import annotations
import io
import sys
import os
import pytest

# Make the project root importable
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from openpyxl import Workbook

from extraction.spir_engine import (
    detect_format, extract_spir,
    OUTPUT_COLS, CI,
    make_new_desc, norm, compute_duplicate_ids,
    detect_spir_type,
)
from extraction.spir_detector import validate_file, detect, ValidationError
from extraction.column_mapper import ColumnMapper, map_sheet, FIELD_KEYWORDS
from services.duplicate_checker import analyse_duplicates
from services.excel_builder import build_xlsx


# ── Helpers ───────────────────────────────────────────────────────────────────

def _wb_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _make_minimal_f2_wb() -> Workbook:
    """Create a minimal FORMAT2-style workbook (single tag, single sheet)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MAIN SHEET"

    # Header structure
    ws["B1"] = "EQUIPMENT  TAG No"
    ws["C1"] = "10-P-1001"          # tag number
    ws["C4"] = "PUMP MODEL X"       # model
    ws["C6"] = "SN-001"             # serial
    ws["C7"] = 1                    # eqpt qty
    ws["Y1"] = "25 SPIR NUMBER:"
    ws["Y2"] = "VEN-0001-TEST"
    ws["C3"] = "Arabian Industries"  # manufacturer
    ws["W3"] = "Arabian Industries"  # supplier
    ws["AB4"] = True                 # checkbox col 28, row 4 = Normal Operating Spares

    # Data header row (row 6)
    ws["G6"] = "ITEM NUMBER"
    ws["H6"] = "TOTAL NO. OF IDENTICAL PARTS FITTED"
    ws["I6"] = "DESCRIPTION OF PARTS"
    ws["Z6"] = "SAP NUMBER"

    # Data rows
    ws["G8"] = 1
    ws["H8"] = 2
    ws["I8"] = "Gasket NPS 6"
    ws["K8"] = "PART-001"
    ws["O8"] = "SupplierA"
    ws["Z8"] = "100001"
    ws["AA8"] = "C - Stock"

    ws["G9"] = 2
    ws["H9"] = 4
    ws["I9"] = "O-Ring Set"
    ws["K9"] = "PART-002"
    ws["O9"] = "SupplierA"
    ws["Z9"] = "100002"
    ws["AA9"] = "C - Stock"

    return wb


# ── Unit tests: helpers ───────────────────────────────────────────────────────

class TestHelpers:
    def test_norm_strips_and_lowercases(self):
        assert norm("  DESCRIPTION OF PARTS  ") == "description of parts"

    def test_norm_collapses_whitespace(self):
        assert norm("DESCRIPTION\n\nOF   PARTS") == "description of parts"

    def test_make_new_desc_full(self):
        result = make_new_desc("Gasket", "PART-001", "SupplierA")
        assert result == "Gasket, PART-001, SupplierA"

    def test_make_new_desc_skips_tba(self):
        result = make_new_desc("Gasket", "TBA", "N/A")
        assert result == "Gasket"

    def test_make_new_desc_skips_na_variants(self):
        for placeholder in ["NA", "N/A", "N.A", "N.A.", "TBC", "-", ".", "NIL",
                             "NONE", "NOT APPLICABLE", "NOT AVAILABLE", "UNKNOWN"]:
            result = make_new_desc("Seal", placeholder, "")
            assert result == "Seal", f"Failed for placeholder: {placeholder}"

    def test_make_new_desc_no_desc(self):
        result = make_new_desc("", "PART-001", "SupplierA")
        assert "PART-001" in result

    def test_make_new_desc_all_empty(self):
        result = make_new_desc("", "", "")
        assert result == ""


# ── Unit tests: duplicate detection ──────────────────────────────────────────

class TestDuplicateDetection:
    def _item(self, desc, part, sap):
        return {"desc": desc, "mfr_part_no": part, "sap_no": sap}

    def test_no_duplicates(self):
        items = [
            self._item("Gasket NPS6", "P001", "100001"),
            self._item("O-Ring Set", "P002", "100002"),
        ]
        labels = compute_duplicate_ids(items)
        assert all(l == "" for l in labels)

    def test_exact_duplicate(self):
        items = [
            self._item("Gasket NPS6", "P001", "100001"),
            self._item("Gasket NPS6", "P001", "100001"),
        ]
        labels = compute_duplicate_ids(items)
        assert labels[0] == labels[1]
        assert labels[0].startswith("Duplicate")

    def test_sap_mismatch(self):
        items = [
            self._item("Gasket NPS6", "P001", "100001"),
            self._item("Gasket NPS6", "P001", "100099"),  # different SAP
        ]
        labels = compute_duplicate_ids(items)
        assert labels[0] == "SAP NUMBER MISMATCH"
        assert labels[1] == "SAP NUMBER MISMATCH"

    def test_duplicate_counter_increments(self):
        items = [
            self._item("ItemA", "P001", "100001"),
            self._item("ItemA", "P001", "100001"),
            self._item("ItemB", "P002", "100002"),
            self._item("ItemB", "P002", "100002"),
        ]
        labels = compute_duplicate_ids(items)
        dup_labels = {l for l in labels if l.startswith("Duplicate")}
        assert len(dup_labels) == 2   # Duplicate 1 and Duplicate 2


# ── Unit tests: SPIR type detection ──────────────────────────────────────────

class TestSpirTypeDetection:
    def _ws_with_checkbox(self, row: int, value) -> object:
        wb = Workbook()
        ws = wb.active
        ws.cell(row, 28).value = value
        return ws

    def test_commissioning(self):
        ws = self._ws_with_checkbox(2, True)
        assert detect_spir_type(ws, 28) == "Commissioning Spares"

    def test_initial(self):
        ws = self._ws_with_checkbox(3, True)
        assert detect_spir_type(ws, 28) == "Initial Spares"

    def test_normal_operating(self):
        ws = self._ws_with_checkbox(4, True)
        assert detect_spir_type(ws, 28) == "Normal Operating Spares"

    def test_life_cycle(self):
        ws = self._ws_with_checkbox(5, True)
        assert detect_spir_type(ws, 28) == "Life Cycle Spares"

    def test_no_checkbox_uses_sheet_name(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Commissioning Spares Sheet"
        assert detect_spir_type(ws, 28) == "Commissioning Spares"


# ── Unit tests: column mapper ─────────────────────────────────────────────────

class TestColumnMapper:
    def _ws_with_headers(self, headers: dict[int, str]) -> object:
        """Create a worksheet with given col→header text in row 6."""
        wb = Workbook()
        ws = wb.active
        for col, text in headers.items():
            ws.cell(6, col).value = text
        return ws

    def test_finds_description(self):
        ws = self._ws_with_headers({9: "DESCRIPTION OF PARTS"})
        mapper = map_sheet(ws, header_rows=[6])
        assert mapper.get("DESCRIPTION") == 9

    def test_finds_sap_number(self):
        ws = self._ws_with_headers({26: "SAP NUMBER"})
        mapper = map_sheet(ws, header_rows=[6])
        assert mapper.get("SAP_NUMBER") == 26

    def test_finds_item_number(self):
        ws = self._ws_with_headers({7: "ITEM NUMBER"})
        mapper = map_sheet(ws, header_rows=[6])
        assert mapper.get("ITEM_NUMBER") == 7

    def test_case_insensitive(self):
        ws = self._ws_with_headers({9: "description of parts"})
        mapper = map_sheet(ws, header_rows=[6])
        assert mapper.get("DESCRIPTION") == 9

    def test_partial_match(self):
        ws = self._ws_with_headers({11: "MANUFACTURER PART NUMBER (SEE NOTE 2)"})
        mapper = map_sheet(ws, header_rows=[6])
        assert mapper.get("MFR_PART_NO") == 11

    def test_missing_returns_none(self):
        ws = self._ws_with_headers({1: "UNRELATED HEADER"})
        mapper = map_sheet(ws, header_rows=[6])
        assert mapper.get("SAP_NUMBER") is None

    def test_coverage(self):
        ws = self._ws_with_headers({
            7: "ITEM NUMBER",
            9: "DESCRIPTION OF PARTS",
            26: "SAP NUMBER",
            11: "MANUFACTURER PART NUMBER",
        })
        mapper = map_sheet(ws, header_rows=[6])
        cov = mapper.coverage()
        assert cov > 0.0

    def test_report_structure(self):
        ws = self._ws_with_headers({9: "DESCRIPTION OF PARTS", 26: "SAP NUMBER"})
        mapper = map_sheet(ws, header_rows=[6])
        report = mapper.report()
        assert "DESCRIPTION" in report
        assert "SAP_NUMBER" in report
        assert report["DESCRIPTION"]["col"] == 9
        assert report["SAP_NUMBER"]["col"] == 26


# ── Unit tests: file validator ────────────────────────────────────────────────

class TestFileValidator:
    def _xlsx_bytes(self) -> bytes:
        wb = Workbook()
        return _wb_bytes(wb)

    def test_valid_xlsx(self):
        data = self._xlsx_bytes()
        validate_file("test.xlsx", data)   # should not raise

    def test_invalid_extension(self):
        with pytest.raises(ValidationError, match="Unsupported"):
            validate_file("test.csv", b"data,data")

    def test_file_too_large(self):
        data = b"x" * (2 * 1024 * 1024)  # 2 MB of garbage
        with pytest.raises(ValidationError):
            validate_file("test.xlsx", data, max_mb=1)

    def test_corrupt_file(self):
        with pytest.raises(ValidationError, match="Could not open"):
            validate_file("test.xlsx", b"not an excel file at all")


# ── Integration: format detection ────────────────────────────────────────────

class TestFormatDetection:
    def test_format2_single_tag(self):
        wb = _make_minimal_f2_wb()
        assert detect_format(wb) in ("FORMAT2", "FORMAT3")

    def test_format1_annexure(self):
        wb = Workbook()
        wb.create_sheet("MAIN SHEET")
        wb.create_sheet("Annexure 1")
        wb.create_sheet("Annexure 2")
        assert detect_format(wb) == "FORMAT1"

    def test_format4_continuation(self):
        wb = Workbook()
        wb.create_sheet("MAIN SHEET")
        wb.create_sheet("Continuation Sheet")
        assert detect_format(wb) == "FORMAT4"

    def test_format5_multi_continuation(self):
        wb = Workbook()
        wb.create_sheet("MAIN SHEET")
        wb.create_sheet("Continuation Sheet(1)")
        wb.create_sheet("Continuation Sheet(2)")
        assert detect_format(wb) == "FORMAT5"


# ── Integration: excel builder ────────────────────────────────────────────────

class TestExcelBuilder:
    def test_builds_valid_xlsx(self):
        rows = [
            ["VEN-001"] + [None] * (len(OUTPUT_COLS) - 1),
        ]
        data = build_xlsx(rows, "VEN-001")
        assert isinstance(data, bytes)
        assert len(data) > 1000   # should be a real xlsx

    def test_output_has_correct_columns(self):
        rows = []
        data = build_xlsx(rows, "VEN-001")
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, len(OUTPUT_COLS) + 1)]
        assert headers == OUTPUT_COLS

    def test_output_has_data_rows(self):
        row = [None] * len(OUTPUT_COLS)
        row[CI["SPIR NO"]]   = "VEN-001"
        row[CI["TAG NO"]]    = "10-P-1001"
        row[CI["ITEM NUMBER"]] = 1
        row[CI["DESCRIPTION OF PARTS"]] = "Test Gasket"
        data = build_xlsx([row], "VEN-001")
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.max_row == 2  # header + 1 data row


# ── Integration: duplicate analyser ──────────────────────────────────────────

class TestDuplicateAnalyser:
    def _row(self, dup_label) -> list:
        row = [None] * len(OUTPUT_COLS)
        row[CI["DUPLICATE ID"]]         = dup_label
        row[CI["TAG NO"]]               = "10-P-001"
        row[CI["ITEM NUMBER"]]          = 1
        row[CI["DESCRIPTION OF PARTS"]] = "Gasket"
        return row

    def test_counts_duplicates(self):
        rows = [self._row("Duplicate 1"), self._row("Duplicate 1"), self._row(0)]
        result = analyse_duplicates(rows)
        assert result["dup1_count"] == 2
        assert result["sap_count"]  == 0

    def test_counts_sap_mismatches(self):
        rows = [self._row("SAP NUMBER MISMATCH"), self._row(0)]
        result = analyse_duplicates(rows)
        assert result["sap_count"]  == 1
        assert result["dup1_count"] == 0

    def test_zero_label_is_not_duplicate(self):
        rows = [self._row(0), self._row(0)]
        result = analyse_duplicates(rows)
        assert result["dup1_count"] == 0
        assert result["sap_count"]  == 0
