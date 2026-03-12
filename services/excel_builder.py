"""
services/excel_builder.py
Builds the styled output Excel workbook from a list of extracted rows.
Completely decoupled from Flask/FastAPI — returns raw bytes.
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from extraction.spir_engine import OUTPUT_COLS


# Column display widths (characters)
_COL_WIDTHS: dict[str, int] = {
    'SPIR NO':                          24,
    'TAG NO':                           22,
    'EQPT MAKE':                        28,
    'EQPT MODEL':                       24,
    'EQPT SR NO':                       12,
    'EQPT QTY':                         10,
    'QUANTITY IDENTICAL PARTS FITTED':  12,
    'ITEM NUMBER':                      10,
    'DESCRIPTION OF PARTS':             50,
    'NEW DESCRIPTION OF PARTS':         60,
    'DWG NO INCL POSN NO':              42,
    'MANUFACTURER PART NUMBER':         36,
    'SUPPLIER OCM NAME':                28,
    'CURRENCY':                         24,
    'UNIT PRICE':                       12,
    'DELIVERY TIME IN WEEKS':           14,
    'MIN MAX STOCK LVLS QTY':           14,
    'UNIT OF MEASURE':                  14,
    'SAP NUMBER':                       16,
    'CLASSIFICATION OF PARTS':          20,
    'DUPLICATE ID':                     22,
    'SHEET':                            22,
    'SPIR TYPE':                        26,
}

# Styles
_HDR_FILL    = PatternFill('solid', fgColor='375623')
_HDR_FONT    = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
_HDR_ALIGN   = Alignment(horizontal='center', vertical='center', wrap_text=True)
_THIN        = Side(style='thin', color='BFBFBF')
_BORDER      = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)
_DATA_FONT   = Font(name='Calibri', size=10, color='000000')
_DATA_ALIGN  = Alignment(vertical='center', wrap_text=False)
_WHITE_FILL  = PatternFill('solid', fgColor='FFFFFF')


def build_xlsx(rows: list[list], spir_no: str = '') -> bytes:
    """
    Build a styled .xlsx workbook from extracted rows.

    Args:
        rows:     List of OUTPUT_COLS-length lists (from extract_spir()['rows']).
        spir_no:  Used only for the return value label; not written to the file.

    Returns:
        Raw bytes of the .xlsx file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SPIR Extraction'

    # ── Header row ────────────────────────────────────────────────────────────
    ws.append(OUTPUT_COLS)
    for cell in ws[1]:
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _HDR_ALIGN
        cell.border    = _BORDER
    ws.row_dimensions[1].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row in rows:
        ws.append(row)
        ri = ws.max_row
        for cell in ws[ri]:
            cell.font      = _DATA_FONT
            cell.fill      = _WHITE_FILL
            cell.border    = _BORDER
            cell.alignment = _DATA_ALIGN
        ws.row_dimensions[ri].height = 15

    # ── Column widths ─────────────────────────────────────────────────────────
    for idx, col_name in enumerate(OUTPUT_COLS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = (
            _COL_WIDTHS.get(col_name, 14)
        )

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
